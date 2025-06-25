
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

st.title("Client Log Matcher")

client_file = st.file_uploader("Upload Client File", type=["csv", "xlsx"])
call_log_file = st.file_uploader("Upload Call Log File", type=["csv", "xlsx"])

def normalize_number(num):
    return ''.join(filter(str.isdigit, str(num)))[-10:]

def get_longest_call_df(call_df):
    call_df["Normalized"] = call_df["Phone Number"].astype(str).str.replace(r"\D", "", regex=True).str[-10:]
    call_df["Talk Time"] = pd.to_numeric(call_df["Talk Time"], errors="coerce").fillna(0)
    return call_df.sort_values("Talk Time", ascending=False).drop_duplicates("Normalized")

if client_file and call_log_file:
    client_df = pd.read_csv(client_file) if client_file.name.endswith(".csv") else pd.read_excel(client_file)
    call_df = pd.read_csv(call_log_file) if call_log_file.name.endswith(".csv") else pd.read_excel(call_log_file)

    # Normalize both phone columns
    client_df["Phone Cleaned"] = client_df.iloc[:, 1].astype(str).apply(normalize_number)
    call_df["Normalized"] = call_df.iloc[:, 5].astype(str).apply(normalize_number)
    call_df["Talk Time"] = pd.to_numeric(call_df["Talk Time"], errors="coerce").fillna(0)

    matched_rows = []
    stats = {}

    for _, row in client_df.iterrows():
        client_num = row["Phone Cleaned"]
        matched = call_df[call_df["Normalized"] == client_num]

        if matched.empty:
            row_data = [row[0], row[1], "NEED TO CHECK", "", ""] + [""] * 5
        else:
            matched_sorted = matched.sort_values("Talk Time", ascending=False)
            top_match = matched_sorted.iloc[0]
            if any(m[1] == client_num for m in matched_rows):
                lead_source = "DUPLICATE"
                vendor = ""
            else:
                lead_source = top_match["Queue Name"]
                vendor = top_match["Source Name"] if lead_source == "MEDICARE CLOSER QUEUE" else ""

            row_data = [row[0], row[1], lead_source, vendor, ""] + [""] * 5
            matched_rows.append((row[0], client_num))

        matched_rows.append(row_data)

    export_df = pd.DataFrame(matched_rows, columns=["Client Name", "Phone Number", "Lead Source", "Vendor", "HEAP DNC", "G", "H", "I", "J", "K"])

    # Stats generation
    lead_sources = export_df["Lead Source"].value_counts()
    stats_df = lead_sources[~lead_sources.index.isin(["DUPLICATE", "NEED TO CHECK"])].reset_index()
    stats_df.columns = ["QUEUE STATS", "SALES"]

    # Add placeholders for calls and abandon
    stats_df["CALLS"] = 0
    stats_df["ABANDON"] = 0
    stats_df["CLOSING %"] = "=IF(H2=0,0,J2/H2)"

    total_row = {
        "QUEUE STATS": "TOTAL",
        "CALLS": 0,
        "ABANDON": 0,
        "SALES": stats_df["SALES"].sum(),
        "CLOSING %": "=IF(H100=0,0,J100/H100)"
    }

    stats_df = pd.concat([stats_df, pd.DataFrame([total_row])], ignore_index=True)

    # Write to Excel with formatting
    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    for r in export_df.values.tolist():
        ws.append(r)

    start_row = len(export_df) + 3
    ws.append([])
    for row in stats_df.itertuples(index=False, name=None):
        ws.append(row)

    headers = ["QUEUE STATS", "CALLS", "ABANDON", "SALES", "CLOSING %"]
    for idx, val in enumerate(headers, start=7):
        ws.cell(row=start_row - 1, column=idx).value = val

    # Borders for the stats table only
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in range(start_row - 1, start_row + len(stats_df)):
        for col in range(7, 12):  # G to K
            ws.cell(row=row, column=col).border = border

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output)
    st.download_button("Download Matched Report", output.getvalue(), file_name="Matched_Client_Stats.xlsx")
    