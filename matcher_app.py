
import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide")
st.title("Client & Call Log Matcher with Stats")

client_file = st.file_uploader("Upload Client List CSV", type=["csv"])
call_file = st.file_uploader("Upload Call Log CSV", type=["csv"])

if client_file and call_file:
    # Always load phone numbers as string to prevent scientific notation
    client_df = pd.read_csv(client_file, dtype={"Phone Number": str})
    call_df = pd.read_csv(call_file)

    # Clean client phone numbers to last 10 digits
    client_df["Phone Number"] = (
        client_df["Phone Number"]
        .astype(str)
        .apply(lambda x: "".join(filter(str.isdigit, x))[-10:] if pd.notna(x) else "")
    )

    # Clean call log phone numbers to last 10 digits
    call_df["Clean Phone"] = (
        call_df["Number Dialed"]
        .astype(str)
        .str.replace(r"\D", "", regex=True)
        .str[-10:]
    )

    # Prepare output
    results = []
    seen_numbers = set()

    for _, row in client_df.iterrows():
        phone = row["Phone Number"]
        matches = call_df[call_df["Clean Phone"] == phone]

        if not phone:
            results.append([row["Client"], phone, "NEED TO CHECK", "", ""])
        elif matches.empty:
            results.append([row["Client"], phone, "NEED TO CHECK", "", ""])
        else:
            top_match = matches.sort_values("Talk Time", ascending=False).iloc[0]
            lead = top_match.iloc[19]
            vendor = top_match.iloc[27] if lead == "MEDICARE CLOSER QUEUE" else ""
            dnc = "X" if "HEAP DNC UPLINE" in str(top_match.iloc[27]) else ""

            if phone in seen_numbers:
                results.append([row["Client"], phone, "DUPLICATE", "", ""])
            else:
                results.append([row["Client"], phone, lead, vendor, dnc])
                seen_numbers.add(phone)

    final_df = pd.DataFrame(results, columns=["Client Name", "Phone Number", "Lead Source", "Vendor", "IFG Suppressed"])

    # Stats table
    counts = final_df["Lead Source"].value_counts()
    queues = [q for q in counts.index if q not in ("DUPLICATE", "NEED TO CHECK")]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Clients"

    # Write client info
    ws.append(["Client Name", "Phone Number", "Lead Source", "Vendor", "IFG Suppressed"])
    for row in final_df.itertuples(index=False):
        ws.append(row)

    # Write stats headers in G2
    stat_headers = ["QUEUE STATS", "CALLS", "ABANDON", "SALES", "CLOSING %"]
    for i, h in enumerate(stat_headers, start=7):
        ws.cell(row=2, column=i, value=h)

    # Write each queue row
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for idx, q in enumerate(queues):
        row_num = idx + 3
        ws.cell(row=row_num, column=7, value=q)
        ws.cell(row=row_num, column=8, value=0)
        ws.cell(row=row_num, column=9, value=0)
        ws.cell(row=row_num, column=10, value=counts[q])
        ws.cell(row=row_num, column=11, value=f"=IF(H{row_num}=0,"",J{row_num}/H{row_num})")
        for c in range(7, 12):
            ws.cell(row=row_num, column=c).border = border
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

    # TOTAL row
    total_row = len(queues) + 3
    ws.cell(row=total_row, column=7, value="TOTAL")
    ws.cell(row=total_row, column=8, value=f"=SUM(H3:H{total_row -1})")
    ws.cell(row=total_row, column=9, value=f"=SUM(I3:I{total_row -1})")
    ws.cell(row=total_row, column=10, value=f"=SUM(J3:J{total_row -1})")
    ws.cell(row=total_row, column=11, value=f"=IF(H{total_row}=0,"",J{total_row}/H{total_row})")
    for c in range(7,12):
        ws.cell(row=total_row, column=c).border = border
        ws.cell(row=total_row, column=c).alignment = Alignment(horizontal="center")

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Output file
    output = io.BytesIO()
    wb.save(output)
    st.download_button("Download Matched Excel", output.getvalue(), file_name="Matched_Client_Stats.xlsx")
